# -*- coding: utf-8 -*-
"""
Created on Feb 9 22:48:40 2020
Revised on Aug 27 2022 by Christopher Pham to utilize the app_password to send message to email from Python
Revised on February 02 2023 By Girija Rijal
Reference: https://www.simplifiedpython.net/python-gui-login/
Reference: https://stackoverflow.com/questions/46268167/how-to-search-for-data-in-an-xlsx-file-using-python-3
Reference: https://www.geeksforgeeks.org/python-simple-registration-form-using-tkinter/?ref=rp
Reference: https://riptutorial.com/tkinter/example/29713/grid--#:~:text=tkinter%20grid()&text=The%20grid()%20geometry%20manager,%2C%20row%20%2C%20rowspan%20and%20sticky%20.
Reference: https://stackoverflow.com/questions/17267140/python-pack-and-grid-methods-together
Reference: https://stackoverflow.com/questions/42491486/setting-an-image-as-a-tkinter-window-background
Reference: https://northernlights.imanet.org/home?ssopc=1
Reference: https://www.freecodecamp.org/news/exception-handling-python/
Reference: https://www.twilio.com/docs/sms/quickstart/python
"""

"""
About this code:

This code is composed of the GUI login from simplifiedpython.net and the Simple Registartion
Form from geeksforgeeks.org (Refernces above). In order to operate both programs from the same
source code each source code for each program had to be chnaged in some manner. For example,
grid and packs cannot be used together in the same code so the code protion from the GUI login
program had to be changed to work with grid. Additionally, the registration form program code had to be
changed to operate as a "pop up screen" upon user login within the GUI code. To make the Registration
form program operate with the GUI interface global vaeiables had to be created for each user input and
related variables. Other program additions include background imagen, which was implimented thanks
to the stackoverflow questions page (referenced above). Other changes include atomatic closure of
windows upon completion of tasks, such as registration form submission. Other chaages also
include popups that notify the user of submission form success and empty text boxes.
"""

# import modules
import yagmail
from tkinter import *
from openpyxl import *
from datetime import date, datetime, timedelta
import pandas as pd
import os
import smtplib, ssl
from twilio.rest import Client


# open existing excel files for user infro
# filesshould be saved in the same directory ot path
wb = load_workbook(r"Simple_Registration_Database.xlsx")
wb2 = load_workbook(r"Registration_UserName_Password.xlsx")


# Create sheets for objects
sheet = wb.active
sheet2 = wb2.active


Twilio_Number="+18334802855", #this is your own Twilio number
account_sid = 'AC679a337***********************a'  #this is your account SID
auth_token = 'c7f01fe6469e**********************'     #this is your own auth_token

from_address = "something@gmail.com" #this is your own gmail account
app_password = 'lkdasndkjanskdjask' # a token for gmail, this is the app password from Gmail Security


# Excel sheet layour for registration form
#   Use to modify Excel file Simple_Registration_Database.xlsx
def excel():

    # formating the excel sheet layout and spacing
    # of columns
    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["C"].width = 35
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 40
    sheet.column_dimensions["G"].width = 50
    sheet.column_dimensions["H"].width = 20
    sheet.column_dimensions["I"].width = 20
    sheet.column_dimensions["J"].width = 20
    sheet.column_dimensions["K"].width = 20
    sheet.column_dimensions["L"].width = 20
    sheet.column_dimensions["M"].width = 20

    sheet.column_dimensions["K"].width = 20
    # Location to write user inputs on excel sheet
    sheet.cell(row=1, column=1).value = "FirstName"
    sheet.cell(row=1, column=2).value = "MiddleName"
    sheet.cell(row=1, column=3).value = "LastName"
    sheet.cell(row=1, column=4).value = "DOB"
    sheet.cell(row=1, column=5).value = "PhoneNumber"
    sheet.cell(row=1, column=6).value = "Email"
    sheet.cell(row=1, column=7).value = "StreetAddress"
    sheet.cell(row=1, column=8).value = "CurrentDate"
    sheet.cell(row=1, column=9).value = "FutureDate"
    sheet.cell(row=1, column=10).value = "Sent"
    sheet.cell(row=1, column=11).value = "SendText"

    sheet.cell(row=1, column=12).value = "Allergies"

    sheet.cell(row=1, column=13).value = "Any Medications"
    sheet.cell(row=1, column=14).value = "Any reminders"
# set focus on the middle_field box
def focus1(event):
    middle_field.focus_set()


# set focus on the last_field box
def focus2(event):
    last_field.focus_set()


# set focus on the form_no_field box
def focus3(event):
    DOB_field.focus_set()


# set focus on the contact_no_field box
def focus4(event):
    phone_field.focus_set()


# set focus on the email_id_field box
def focus5(event):
    email_field.focus_set()


# set focus on the address_field box
def focus6(event):
    address_field.focus_set()


def focus7(event):
    current_date_field.focus_set()


def focus8(event):
    Allergies_field.focus_set()
    
def focus9(event):
     Fathername_field.focus_set()
     
def focus10(event):
    Mothername_field.focus_set()

# call function to clear text entry boxes
def clear():

    # clear the content of text entry box
    name_field.delete(0, END)
    middle_field.delete(0, END)
    last_field.delete(0, END)
    DOB_field.delete(0, END)
    phone_field.delete(0, END)
    email_field.delete(0, END)
    address_field.delete(0, END)
    current_date_field.delete(0, END)
    current_date_field.delete(0, END)
    Allergies_field.delete(0,END)
    Fathername_field.delete(0,END)
    Mothername_field.delete(0,END)
# Empty input screen to notify user of empty text boxes
# If any one text box is empty is empty the form will not submit and
# the user wilbe notifies with a pop up screen
def Empty_Input():
    global empty_input_screen
    # Keep empty input popup ontop of current windows
    empty_input_screen = Toplevel()
    empty_input_screen.title("Error")
    empty_input_screen.geometry("200x100")
    # Display empty text entry to user
    Label(empty_input_screen, text="Empty text entry!", font=("Calibri", 13)).grid(
        row=0, column=0, padx=25
    )
    # ok to exit and return to registration form
    Button(
        empty_input_screen,
        text="OK to exit",
        width=15,
        height=1,
        bg="red",
        font=("Calibri", 13),
        command=empty_input_error,
    ).grid(row=1, column=0, padx=25)


# Notify the user of a susseccfull registration form submission_success
# Close out registration form upon pressing ok to exit button
def submission_success():
    global submission_success_screen
    # keep submittion popup ontop
    submission_success_screen = Toplevel()
    submission_success_screen.title("Submission Success")
    submission_success_screen.geometry("200x100")
    # display to user submission success and exit form
    Label(
        submission_success_screen, text="submission Success!", font=("Calibri", 13)
    ).grid(row=0, column=0, padx=25)
    Button(
        submission_success_screen,
        text="OK to exit",
        width=15,
        height=1,
        bg="red",
        font=("Calibri", 13),
        command=delete_submission_success_screen,
    ).grid(row=1, column=0, padx=25)


# Notify the user of a susseccfull input for professor date input test tool
def submission_success2():
    global submission_success_screen2

    # keep submittion popup ontop
    submission_success_screen2 = Toplevel()
    submission_success_screen2.title("Submission Success")
    submission_success_screen2.geometry("200x100")
    # display to user submission success and exit form
    Label(
        submission_success_screen2, text="submission Success!", font=("Calibri", 13)
    ).grid(row=0, column=0, padx=25)
    Button(
        submission_success_screen2,
        text="OK to exit",
        width=15,
        height=1,
        bg="red",
        font=("Calibri", 13),
        command=delete_submission_success_screen2,
    ).grid(row=1, column=0, padx=25)


# Design the date check popup tool screen
def current_date_check():
    global date_check_screen
    date_check_screen = Toplevel()
    global date_check_entry
    global date_check_in
    date_check_in = StringVar()
    # keep submittion popup ontop
    date_check_screen.title("Date Check Screen (Developer Testing Tool)")
    date_check_screen.geometry("800x180")

    # Display label with the first vaccination date and second vaccination date
    display_submission()
    Label(
        date_check_screen,
        text="Submit 2nd due date to check that the reminder is working (format: YYYY-MM-DD):",
        font=("Calibri", 13),
    ).grid(row=0, column=0, padx=25)
    date_check_entry = Entry(date_check_screen, textvariable=date_check_in)
    date_check_entry.grid(row=0, column=1)

    Button(
        date_check_screen,
        text="Submit",
        width=15,
        height=1,
        bg="red",
        font=("Calibri", 13),
        command=date_verify,
    ).grid(row=3, column=1, padx=30)


# Verify that a second email will be sent to recipient that just registered
# 21 days from the first vaccine date
def first_vaccine():
    #read dataframe df
    df = pd.read_csv("COVID_Vaccine_Database.csv")

    # initiate text messaging sever from twilio (Reference listed above importe modules)
    client = Client(
        account_sid, auth_token  # SID & Auth Token
    )
    rows = df.shape[0]

    # format the dates to be sent to user with strptime and strftime
    first_vac = df.iloc[rows - 1, 7]
    first_vac = datetime.strptime(first_vac, "%Y-%m-%d %H:%M:%S")
    first_vac = first_vac.strftime("%B %d, %Y")
    sec_vac = df.iloc[rows - 1, 8]
    sec_vac = datetime.strptime(sec_vac, "%Y-%m-%d %H:%M:%S")
    #sec_vac = datetime.strptime(sec_vac, "%Y-%m-%d")
    sec_vac = sec_vac.strftime("%B %d, %Y")

    Email = df.iloc[rows - 1, 5]
    LastName = df.iloc[rows - 1, 2]
    FirstName = df.iloc[rows - 1, 0]
    text = df.iloc[rows - 1, 10]


    #####################################
    #The following code sends email out #
    #####################################

    message = """Subject: First COVID-19 Vaccination Completed

    Hello {First} {Last} you received your first COVID-19 Vaccination today on {Vaccination1} \n

    Your second vaccination will be on {Vaccination2}.
    """


    to_address = Email  #send test to another email or the same email is OK

    subject = 'First COVID-19 Vaccination Completed'  #modify the subject line anyway you like
    email_content = message.format(First=FirstName, Last=LastName, Vaccination1=first_vac,Vaccination2=sec_vac) #refer to message format above

    with yagmail.SMTP(from_address, app_password) as yag:
        yag.send(to_address, subject, email_content)
        print('Sent 1st Covid vaccination email successfully to {} {}'.format(FirstName, LastName))  #you can have different success message



    ###########################################
    #The following code sends text message out#
    ###########################################


    client = Client(account_sid, auth_token)


    message = client.messages.create(
                        body=message.format(
                            First=FirstName,
                            Last=LastName,
                            Vaccination1=first_vac,
                            Vaccination2=sec_vac,
                            ),
                        from_=Twilio_Number, #this is your own Twilio number
                        to=text,  #this is the patient's phone number to receive the text msg, use to=round(text) if Twilio appends a 0 to your phone number.
                        )

    print(message.sid)
    print('Sent 1st Covid vaccination text successfully to {0} {1}'.format(FirstName, LastName))  #you can have different success message
    print("----------------")





# Second vaccine notification function
def second_vaccine():
    #read dataframe df
    df = pd.read_csv("COVID_Vaccine_Database.csv")

    # initiate text messaging sever from twilio (Reference listed above importe modules)
    client = Client(
        account_sid, auth_token
    )



    # time delta of 3 days and datetime.now() for current time
    # helps check the current date and second vaccination date
    check = timedelta(days=3)
    check = check.days
    now = datetime.now()

    # number of rows in datframes df (COVID_Vaccine_Database)
    rows = df.shape[0]
    # second vaccination date
    sec_vac = df.iloc[rows - 1, 8]

    Email = df.iloc[rows - 1, 5]
    LastName = df.iloc[rows - 1, 2]
    FirstName = df.iloc[rows - 1, 0]
    Text = df.iloc[rows - 1, 10]


    to_address = Email  #send test to another email or the same email is OK


    #Scan all rows in the Excel file for rows missing reminding text messages, check for date < 3 days and send the email and text message
    for i, r in df.iterrows():
        FD = datetime.strptime(r["FutureDate"], "%Y-%m-%d %H:%M:%S")
        FD2 = FD.strftime("%B %d, %Y")
        # check to see if the current date is 3days or less from second vaccination date
        if (
            abs(datetime.strptime(r["FutureDate"], "%Y-%m-%d %H:%M:%S") - now).days
            <= check
        ):

            # check to see if a seccond vaccination notifaction has been sent already
            # if message has not been sent the value is 1, then send the sendond vaccination date and update the sent value to 0.
            if r["Sent"] == 1:

                #####################################
                #The following code sends email out #
                #####################################
                subject = 'Reminder for 2nd vaccination'  #modify the subject line anyway you like
                message = """Subject: Reminder for Second COVID-19 Vaccination

                Hello {First} {Last} your second COVID-19 vaccination is coming up on {Vaccination2}."""

                email_content = message.format(First=r["FirstName"], Last=r["LastName"], Vaccination2=FD2) #refer to message format above

                with yagmail.SMTP(from_address, app_password) as yag:
                    yag.send(to_address, subject, email_content)
                    print('Sent 2nd Covid vaccination email successfully to {0} {1}'.format(r["FirstName"], r["LastName"]))  #you can have different success message


                ###########################################
                #The following code sends text message out#
                ###########################################


                client = Client(account_sid, auth_token)


                message = client.messages.create(
                                    body=message.format(
                                        First=r["FirstName"],
                                        Last=r["LastName"],
                                        Vaccination2=FD2,
                                        ),
                                    from_=Twilio_Number, #this is your own Twilio number
                                    to="+14156406283", #this is the patient's phone number to receive the text msg, use to=round(text) if Twilio appends a 0 to your phone number.
                                    )

                print(message.sid)
                print('Sent 2nd Covid vaccination text successfully to {0} {1}', format(r["FirstName"], r["LastName"]))  #you can have different success message
                print("----------------")

                # update database, excel file and csv
                df.loc[i, "Sent"] = 0
                df.to_excel("Simple_Registration_Database.xlsx", index=False)
                df.to_csv("COVID_Vaccine_Database.csv", index=None, header=True)


# verify the date to be less than three days from second vaccination for the
# professor tool
def date_verify():
    date_input = datetime.strptime(date_check_entry.get(), "%Y-%m-%d") + timedelta(
        minutes=5
    )
    submission_success2()
    df = pd.read_csv("COVID_Vaccine_Database.csv")
    # start text messaging server from twilio account
    client = Client(
        account_sid, auth_token
    )
    # time delta of 3 days
    check = timedelta(days=3)
    check = check.days
    rows = df.shape[0]

    # second vaccination date from dataframe
    # obtaine values from datframe through their corresponding index Location
    sec_vac = df.iloc[rows - 1, 8]
    sec_vac2 = datetime.strptime(sec_vac, "%Y-%m-%d %H:%M:%S") + timedelta(minutes=5)
    sec_vac_send = sec_vac2.strftime("%B %d, %Y") #keep only date & strip out the time

    Email = df.iloc[rows - 1, 5]
    LastName = df.iloc[rows - 1, 2]
    FirstName = df.iloc[rows - 1, 0]
    Text = df.iloc[rows - 1, 10]

    #####################################
    #The following code sends email out #
    #####################################
    subject = 'Reminder for 2nd vaccination'  #modify the subject line anyway you like
    message = """Subject: Reminder for Second COVID-19 Vaccination

    Hello {First} {Last} your second COVID-19 vaccination is coming up on {Vaccination2}."""

    email_content = message.format(First=FirstName, Last=LastName, Vaccination2=sec_vac_send) #refer to message format above

    to_address = Email  #send test to another email or the same email is OK

    with yagmail.SMTP(from_address, app_password) as yag:
        yag.send(to_address, subject, email_content)
        print('Date-Verify: Sent 2nd Covid vaccination email successfully to {0} {1}'.format( FirstName, LastName))  #you can have different success message


    ###########################################
    #The following code sends text message out#
    ###########################################


    client = Client(account_sid, auth_token)


    message = client.messages.create(
                        body=message.format(
                            First=FirstName,
                            Last=LastName,
                            Vaccination2=sec_vac_send, #note: to show only date, not time
                            ),
                        from_=Twilio_Number, #this is your own Twilio number
                        to=Text, #this is the patient's phone number to receive the text msg, use to=round(text) if Twilio appends a 0 to your phone number.
                        )

    print(message.sid)
    print('Date-Verify: Sent 2nd Covid vaccination text successfully to {0} {1}'.format(FirstName, LastName))  #you can have different success message



# Write unser input into excel sheet2
# if either text entry box is empty the user will be notified
def insert():
    # if user not fill any or one entry box
    # then pop up empty input screen
    if (
        name_field.get() == ""
        or last_field.get() == ""
        or DOB_field.get() == ""
        or phone_field.get() == ""
        or email_field.get() == ""
        or address_field.get() == ""
        or current_date_field.get() == ""
        
        or Allergies_field.get() == ""
        or Fathername_field.get() == ""
        or Mothername_field.get() == ""
        
        
    ):
        Empty_Input()

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # Recive uer input from text entry box and submit to excel sheet
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = middle_field.get()
        sheet.cell(row=current_row + 1, column=3).value = last_field.get()
        sheet.cell(row=current_row + 1, column=4).value = DOB_field.get()
        sheet.cell(row=current_row + 1, column=5).value = phone_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # Recive date input and add a 5 inute timedelta to avoid datetime errors
        sheet.cell(row=current_row + 1, column=8).value = datetime.strptime(
            current_date_field.get(), "%Y-%m-%d"
        ) + timedelta(minutes=5)

        # Add 21 days yo the original date input
        sheet.cell(row=current_row + 1, column=9).value = datetime.strptime(
            current_date_field.get(), "%Y-%m-%d"
        ) + timedelta(days=21, minutes=5)
        sheet.cell(row=current_row + 1, column=10).value = 1
        # add +1 to phone number input
        sheet.cell(row=current_row + 1, column=11).value = "+1" + phone_field.get()
        # sheet.cell(row=current_row + 1, column=9).value = future_date_field.get()
        sheet.cell(row=current_row + 1, column=12).value = Allergies_field.get()
        sheet.cell(row=current_row + 1, column=13).value = Fathername_field.get()
        sheet.cell(row=current_row + 1, column=14).value = Mothername_field.get()
        # save the file
        wb.save(r"Simple_Registration_Database.xlsx")

        df_csv = pd.read_excel(r"Simple_Registration_Database.xlsx")
        df_csv.to_csv(r"COVID_Vaccine_Database.csv", index=None, header=True)

        # set focus on the name_field box
        name_field.focus_set()
        clear()
        submission_success(), first_vaccine(),
        second_vaccine(), current_date_check()


# impliment user registry form window for user registration
def register_form():
    global root
    # keep screen on top
    root = Toplevel()
    root.title("Registration Form")
    root.geometry("1000x600")

    # set background image to registration form main_screen
    # Reference: https://stackoverflow.com/questions/42491486/setting-an-image-as-a-tkinter-window-background
    # import image and sst as backgroung
    bg = PhotoImage(file=r"five.png", height=400, width=1000)
    bg_label = Label(root, image=bg)
    bg_label.image = bg  # keep a reference!
    bg_label.grid(row=0, column=0, columnspan=20, rowspan=20)

    # create global variables for registration form inputs to successfully
    # implimnet with GUI program across the rest of the program
    global Name_in
    global course_in
    global sem_in
    global grad_in
    global phone_in
    global email_in
    global address_in
    global current_date_in
    global Allergies_in
    global Fathername_in
    global Mothername_in

    global name_field
    global middle_field
    global last_field
    global DOB_field
    global phone_field
    global email_field
    global address_field
    global current_date_field
    global Allergies_field
    global Fathername_field
    global Mothername_field
    # Widget text variables
    # https://stackoverflow.com/questions/51783852/what-is-the-difference-between-a-variable-and-stringvar-of-tkinter/51785046
    Name_in = StringVar()
    course_in = StringVar()
    sem_in = StringVar()
    grad_in = StringVar()
    phone_in = StringVar()
    email_in = StringVar()
    address_in = StringVar()
    current_date_in = StringVar()
    Allergies_in = StringVar()
    Fathername_in = StringVar()
    Mothername_in = StringVar()

    excel()

    # create a Form label
    heading = Label(
        root,
        text="COVID-19 Vaccination Form",
        fg="white",
        bg="black",
        font=("Calibri", 13),
    )

    # create a Name label
    name = Label(root, text="First Name", bg="black", fg="red", font=("Calibri", 13))

    # create a Course label
    course = Label(
        root, text="Middle Name", bg="black", fg="red", font=("Calibri", 13)
    )

    # create a Semester label
    sem = Label(root, text="Last Name", bg="black", fg="red", font=("Calibri", 13))

    # create a graduation year. lable
    grad = Label(root, text="DOB", bg="black", fg="red", font=("Calibri", 13))

    # create a phone No. label
    phone = Label(
        root,
        text="10 digit Phone No. (0123456789) ",
        bg="black",
        fg="red",
        font=("Calibri", 13),
    )

    # create a Email label
    email = Label(root, text="Email", bg="black", fg="red", font=("Calibri", 13))

    # create a address label
    address = Label(
        root, text="Street Address", bg="black", fg="red", font=("Calibri", 13)
    )

    # create a current date label
    c_date = Label(
        root,
        text="Current Date (YYYY-MM-DD)",
        bg="black",
        fg="red",
        font=("Calibri", 13),
    )
    Allergies = Label(root, text="Allergies", bg="black", fg="red", font=("Calibri", 13))
    Fathername = Label(root, text="Fathername", bg="black", fg="red", font=("Calibri", 13))
    Mothername = Label(root, text="Mothername", bg="black", fg="red", font=("Calibri", 13))


    # headings postion on the scree with the use of grid.
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    grad.grid(row=4, column=0)
    phone.grid(row=5, column=0)
    email.grid(row=6, column=0)
    address.grid(row=7, column=0)
    c_date.grid(row=8, column=0)
    Allergies.grid(row=9, column=0)
    Fathername.grid(row=10, column=0)
    Mothername.grid(row=11, column=0)

    name_field = Entry(root, textvariable=Name_in)
    middle_field = Entry(root, textvariable=course_in)
    last_field = Entry(root, textvariable=sem_in)
    DOB_field = Entry(root, textvariable=grad_in)
    phone_field = Entry(root, textvariable=phone_in)
    email_field = Entry(root, textvariable=email_in)
    address_field = Entry(root, textvariable=address_in)
    current_date_field = Entry(root, textvariable=current_date_in)
    Allergies_field = Entry(root, textvariable=Allergies_in)
    Fathername_field = Entry(root, textvariable=Fathername_in)
    Mothername_field = Entry(root, textvariable=Mothername_in)

    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    middle_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    last_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    DOB_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    phone_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    email_field.bind("<Return>", focus6)

    address_field.bind("<Return>", focus7)
    
    Allergies_field.bind("<Return>",focus8)
    
    Fathername_field.bind("<Return>",focus9)
    
    Mothername_field.bind("<Return>",focus10)

    # current_date_field.bind("<Return>", focus7)

    # position text entry boxes with grid function in their respected rows/columns
    # column 1 is used ofr text boxes for user input
    name_field.grid(row=1, column=1, ipadx="100")
    middle_field.grid(row=2, column=1, ipadx="100")
    last_field.grid(row=3, column=1, ipadx="100")
    DOB_field.grid(row=4, column=1, ipadx="100")
    phone_field.grid(row=5, column=1, ipadx="100")
    email_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")
    current_date_field.grid(row=8, column=1, ipadx="100")
    Allergies_field.grid(row=9, column=1, ipadx="100")
    Fathername_field.grid(row=10, column=1, ipadx="100")
    Mothername_field.grid(row=11, column=1, ipadx="100")
    # submission button with red background and font change i sused to submit user information
    # into excel file
    submit = Button(
        root,
        text="Submit",
        fg="Black",
        relief="ridge",
        bg="grey",
        width=15,
        height=1,
        font=("Calibri", 15),
        command=insert,
    )
    submit.grid(row=13, column=1)


# design register screen
def register():
    global register_screen
    register_screen = Toplevel(main_screen)
    register_screen.title("Register")
    register_screen.geometry("300x420")

    # Background image for window
    # refrenced above
    bg = PhotoImage(file=r"one.png", height=800, width=800)
    bg_label = Label(register_screen, image=bg)
    bg_label.image = bg  # keep a reference!
    bg_label.grid(row=0, column=0, columnspan=20, rowspan=20)

    global username
    global password
    global username_entry
    global password_entry
    username = StringVar()
    password = StringVar()

    # labe displaied for users to submit username and password information
    Label(
        register_screen,
        text="Please enter details below to register",
        fg="white",
        bg="black",
        width="30",
        height="2",
        font=("Calibri", 13),
    ).grid(row=0, column=0, padx="10")

    # labels for user entries and text entries
    # buttons changees include font color and background color
    username_lable = Label(
        register_screen,
        text="Username * ",
        fg="white",
        bg="black",
        font=("Calibri", 13),
    )
    username_lable.grid(row=2, column=0)
    username_entry = Entry(register_screen, textvariable=username)
    username_entry.grid(row=3, column=0)
    password_lable = Label(
        register_screen,
        text="Password * ",
        fg="white",
        bg="black",
        font=("Calibri", 13),
    )
    password_lable.grid(row=4, column=0)
    password_entry = Entry(register_screen, textvariable=password, show="*")
    password_entry.grid(row=5, column=0)

    Button(
        register_screen,
        text="Register",
        width=15,
        height=1,
        bg="grey",
        font=("Calibri", 13),
        command=register_user,
    ).grid(row=7, column=0)


# Designing window for login
def login():
    global login_screen
    login_screen = Toplevel(main_screen)
    login_screen.title("Login")
    login_screen.geometry("320x400")

    bg = PhotoImage(file=r"four.png", height=400, width=320)
    bg_label = Label(login_screen, image=bg)
    bg_label.image = bg  # keep a reference!
    bg_label.grid(row=0, column=0, columnspan=20, rowspan=20)

    Label(
        login_screen,
        text="Please enter details below to login",
        fg="white",
        bg="black",
        width="30",
        height="2",
        font=("Calibri", 13),
    ).grid(row=0, column=0, padx="10")

    global username_verify
    global password_verify

    username_verify = StringVar()
    password_verify = StringVar()

    global username_login_entry
    global password_login_entry

    Label(
        login_screen, text="Username * ", fg="white", bg="black", font=("Calibri", 13)
    ).grid(row=3, column=0)
    username_login_entry = Entry(login_screen, textvariable=username_verify)
    username_login_entry.grid(row=4, column=0)

    Label(
        login_screen, text="Password * ", fg="white", bg="black", font=("Calibri", 13)
    ).grid(row=6, column=0)
    password_login_entry = Entry(login_screen, textvariable=password_verify, show="*")
    password_login_entry.grid(row=7, column=0)

    Button(
        login_screen,
        text="Login",
        width=15,
        height=1,
        bg="grey",
        font=("Calibri", 13),
        command=login_verify,
    ).grid(row=9, column=0)


# Write user name and password to spreadhseet
#   Use to modify Excel file
# Registration_UserName_Password.xlsx
def excel2():

    # resize the width of columns in
    # excel spreadsheet
    sheet2.column_dimensions["A"].width = 30
    sheet2.column_dimensions["B"].width = 10

    # write given data to an excel spreadsheet
    # at particular location
    # Columns for user name and password
    sheet2.cell(row=1, column=1).value = "User Name"
    sheet2.cell(row=1, column=2).value = "Password"


# Implementing event on register button
def register_user():
    username_info = username.get()
    password_info = password.get()
    excel2()
    current_row2 = sheet2.max_row
    current_column2 = sheet2.max_column

    sheet2.cell(row=current_row2 + 1, column=1).value = username_info
    sheet2.cell(row=current_row2 + 1, column=2).value = password_info

    wb2.save("./Registration_UserName_Password.xlsx")

    username_entry.delete(0, END)
    password_entry.delete(0, END)

    Label(
        register_screen,
        text="Registration Success",
        bg="black",
        fg="red",
        font=("calibri", 11),
    ).grid(row=10, column=0)


def display_submission():
    df = pd.read_csv("COVID_Vaccine_Database.csv")
    rows = df.shape[0]
    first_vac = df.iloc[rows - 1, 7]
    sec_vac = df.iloc[rows - 1, 8]
    FV = datetime.strptime(first_vac, "%Y-%m-%d %H:%M:%S")
    #FV = datetime.strptime(first_vac, "%Y-%m-%d")
    FV2 = FV.strftime("%Y-%m-%d")
    FV3 = datetime.strptime(sec_vac, "%Y-%m-%d %H:%M:%S")
    #FV3 = datetime.strptime(sec_vac, "%Y-%m-%d")
    FV4 = FV3.strftime("%Y-%m-%d")
    message = """You submitted {Vaccination1} for the first vaccination date
    The 2nd Vaccination date is on {Vaccination2}.
    To test whether you will receive a 2nd reminder,
    enter a date in the format YYYY-MM-DD within
    3 days before the second vaccination date"""

    message2 = message.format(Vaccination1=FV2, Vaccination2=FV4)
    Label(
        date_check_screen, text=message2, bg="black", fg="red", font=("calibri", 11)
    ).grid(row=1, column=0)


# Implementing event on login button
def login_verify():
    username1 = username_verify.get()
    password1 = password_verify.get()
    username_login_entry.delete(0, END)
    password_login_entry.delete(0, END)

    # create name list to search from for login dat
    # create passwords list to search from
    # https://stackoverflow.com/questions/51800122/using-openpyxl-to-find-rows-that-contain-cell-with-specific-value-python-3-6
    names = []
    passwords = []

    for i in range(2, sheet2.max_row + 1):
        if sheet2[i][0].value:
            names.append(sheet2[i][0].value)
            passwords.append(sheet2[i][1].value)

    # exempt ValueError when user inputs a username not in file
    # https://www.freecodecamp.org/news/exception-handling-python/
    try:
        index = names.index(username1)
    except ValueError:
        print("")

    # if password and names are saved allow user to register_form
    # otherwize no user found or incorrect password
    # Check to see if the index number of the username matches with its corresponding user password
    if username1 in names:
        if password1 == passwords[index]:
            login_sucess()
            register_form()
        else:
            password_not_recognised()
    else:
        user_not_found()


# Designing popup for login succes
def login_sucess():
    global login_success_screen
    login_success_screen = Toplevel(login_screen)
    login_success_screen.title("Success")
    login_success_screen.geometry("200x100")

    # display Login success to user
    Label(login_success_screen, text="Login Success!", font=("Calibri", 13)).grid(
        row=0, column=0, padx=25
    )

    # close login success window when pressing button
    Button(
        login_success_screen,
        text="OK to exit",
        width=15,
        height=1,
        bg="red",
        font=("Calibri", 13),
        command=delete_login_success,
    ).grid(row=1, column=0, padx=25)


# Designing popup for login invalid password
def password_not_recognised():
    global password_not_recog_screen
    password_not_recog_screen = Toplevel(login_screen)
    password_not_recog_screen.title("Error")
    password_not_recog_screen.geometry("200x100")

    # notify user of invalid password
    Label(
        password_not_recog_screen, text="Invalid Password ", font=("Calibri", 13)
    ).grid(row=0, column=0, padx=25)
    Button(
        password_not_recog_screen,
        text="OK to exit",
        width=15,
        height=1,
        bg="red",
        font=("Calibri", 13),
        command=delete_password_not_recognised,
    ).grid(row=1, column=0, padx=25)


# Designing popup for user not found
def user_not_found():
    global user_not_found_screen
    user_not_found_screen = Toplevel(login_screen)
    user_not_found_screen.title("Error")
    user_not_found_screen.geometry("200x100")

    # notify user of invalid username
    Label(user_not_found_screen, text="User Not Found", font=("Calibri", 13)).grid(
        row=0, column=0
    )
    Button(
        user_not_found_screen,
        text="OK to exit",
        width=15,
        height=1,
        bg="red",
        font=("Calibri", 13),
        command=delete_user_not_found_screen,
    ).grid(row=1, column=0, padx=25)


# Deleting popups
# Delete login success screen
def delete_login_success():
    login_success_screen.destroy()
    login_screen.destroy()


# Delete empty input screen
def empty_input_error():
    empty_input_screen.destroy()


# Delete submission success screen and registration form screen
def delete_submission_success_screen():
    submission_success_screen.destroy()


def delete_submission_success_screen2():
    submission_success_screen2.destroy()
    date_check_screen.destroy()


# Delete password not recognized screen
def delete_password_not_recognised():
    password_not_recog_screen.destroy()


# Delete uset not found screen
def delete_user_not_found_screen():
    user_not_found_screen.destroy()


# delete login screen
def delete_login_screen():
    login_screen.destroy()


# Designing Main(first) window
def main_account_screen():
    global main_screen
    global df

    # Generate GUI
    main_screen = Tk()

    # Resize main screen
    main_screen.geometry("800x400")
    main_screen.title("COVID Vaccination Login Window")

    # Import and set ground of main screen
    bg = PhotoImage(file=r"two.png", height=400, width=800)
    bg_label = Label(main_screen, image=bg)
    bg_label.image = bg  # keep a reference!
    bg_label.grid(row=0, column=0, columnspan=20, rowspan=20)

    Label(
        text="Welcome to EE104 COVID-19 Vaccination Registration Portal!",
        fg="gold",
        bg="#000000", #all RGB = 0000FF
        width="57",
        height="3",
        font=("Calibri", 20),
    ).grid(row=0, column=2, padx=50)

    Label(
        main_screen,
        relief="solid",
        text="Press a button below",
        width="25",
        height="1",
        fg="white",
        bg="red",
        font=("Calibri", 13),
    ).grid(row=1, column=2, padx=50)

    Button(text="Login", height="2", width="30", bg="orange", command=login).grid(
        row=3, column=2
    )

    Button(text="Register", height="2", width="30", bg="grey", command=register).grid(
        row=5, column=2
    )

    main_screen.mainloop()


main_account_screen()
